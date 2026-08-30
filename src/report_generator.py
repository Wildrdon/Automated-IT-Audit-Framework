import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint

def create_excel_workpaper(output_filename="IT_Audit_Workpaper.xlsx"):
    data = [
        {
            "Kullanıcı Adı": "hacker_test",
            "Son Giriş Tarihi": "2026-08-27",
            "Risk Seviyesi": "Kritik",
            "ISO 27001 Maddesi": "A.9.2.3 (Yetkili Erişim)",
            "Bulgu": "Hesap UID 0 (Root) yetkisiyle yetkisiz oluşturulmuş.",
            "Öneri": "Hesap derhal silinmeli ve oluşturan yetkili incelenmeli."
        },
        {
            "Kullanıcı Adı": "nopass_user",
            "Son Giriş Tarihi": "2026-08-27",
            "Risk Seviyesi": "Kritik",
            "ISO 27001 Maddesi": "A.9.4.3 (Parola Yönetim)",
            "Bulgu": "Aktif oturum açma yetkili hesabın parolası boş.",
            "Öneri": "Güçlü parola politikası atanmalı veya hesap kapatılmalı."
        },
        {
            "Kullanıcı Adı": "S-1-5-18",
            "Son Giriş Tarihi": "2026-08-28",
            "Risk Seviyesi": "Orta",
            "ISO 27001 Maddesi": "A.12.4.1 (Olay Günlükleme)",
            "Bulgu": "10 dakika içinde 8 başarısız giriş denemesi (4625).",
            "Öneri": "Servis hesabı parolası ve bağımlı servisler kontrol edilmeli."
        },
        {
            "Kullanıcı Adı": "defaultuser0",
            "Son Giriş Tarihi": "2026-08-27",
            "Risk Seviyesi": "Düşük",
            "ISO 27001 Maddesi": "A.9.2.6 (Erişim Haklarının Kaldırılması)",
            "Bulgu": "Varsayılan kurulum hesabı sistemde pasif duruyor.",
            "Öneri": "Gerekli görülmüyorsa hesap devre dışı bırakılmalı."
        }
    ]

    df = pd.DataFrame(data)
    wb = Workbook()
    
    # ---------------------------------------------------------
    # SHEET 1: EXECUTIVE SUMMARY
    # ---------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Stiller
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    white_bold_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    ws_summary["B2"] = "IT DENETİM YÖNETİCİ ÖZETİ"
    ws_summary["B2"].font = title_font

    ws_summary["B4"] = "Metrik"
    ws_summary["C4"] = "Değer"
    ws_summary["B4"].fill = navy_fill
    ws_summary["C4"].fill = navy_fill
    ws_summary["B4"].font = white_bold_font
    ws_summary["C4"].font = white_bold_font
    ws_summary["B4"].alignment = Alignment(horizontal="left", vertical="center")
    ws_summary["C4"].alignment = Alignment(horizontal="center", vertical="center")

    summary_rows = [
        ("Toplam Taranan Olay", 19223),
        ("Kritik Risk Sayısı", 2),
        ("Orta Risk Sayısı", 1),
        ("Düşük Risk Sayısı", 1)
    ]

    for idx, (metrik, deger) in enumerate(summary_rows, start=5):
        ws_summary[f"B{idx}"] = metrik
        ws_summary[f"C{idx}"] = deger
        ws_summary[f"B{idx}"].font = bold_font if idx > 5 else regular_font
        ws_summary[f"C{idx}"].font = bold_font if idx > 5 else regular_font
        ws_summary[f"B{idx}"].border = thin_border
        ws_summary[f"C{idx}"].border = thin_border
        ws_summary[f"C{idx}"].alignment = Alignment(horizontal="center", vertical="center")
        if idx == 5:
            ws_summary[f"C{idx}"].number_format = '#,##0'

    # Sütun Genişlikleri Ayarı
    ws_summary.column_dimensions['A'].width = 3
    ws_summary.column_dimensions['B'].width = 28
    ws_summary.column_dimensions['C'].width = 15

    # Pasta Grafiği ve Özel Renklendirme
    pie = PieChart()
    labels = Reference(ws_summary, min_col=2, min_row=6, max_row=8)
    chart_data = Reference(ws_summary, min_col=3, min_row=5, max_row=8)
    pie.add_data(chart_data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Risk Dağılım Grafiği"
    pie.width = 15
    pie.height = 7.5

    slice_colors = ["FF5C5C", "FFC000", "92D050"]
    series = pie.series[0]
    for i, color in enumerate(slice_colors):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = color
        series.data_points.append(dp)

    ws_summary.add_chart(pie, "E4")

    # ---------------------------------------------------------
    # SHEET 2: DETAILED FINDINGS
    # ---------------------------------------------------------
    ws_details = wb.create_sheet(title="Detailed Findings")
    ws_details.views.sheetView[0].showGridLines = True

    for r in dataframe_to_rows(df, index=False, header=True):
        ws_details.append(r)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for cell in ws_details[1]:
        cell.fill = navy_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws_details.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=6):
        risk_val = row[2].value
        for cell in row:
            cell.border = thin_border
            cell.font = regular_font
            if risk_val == "Kritik":
                cell.fill = red_fill
            elif risk_val == "Orta":
                cell.fill = yellow_fill

    for col in ws_details.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws_details.column_dimensions[col_letter].width = max(max_len + 4, 15)

    wb.save(output_filename)
    print(f"[+] Şık Excel Çalışma Kağıdı Oluşturuldu: {output_filename}")

if __name__ == "__main__":
    create_excel_workpaper()
